import io
import re
import urllib.error

import openpyxl

from netfetch import urlopen

SHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")

# Жылдық отчеттағы ай-топ атаулары нақты потоктың кодына сәйкес келеді —
# бұл сәйкестік тұрақты (жыл сайын өзгермейді): әр поток әрдайым сол бір
# айда ашылады. 'ШІЛДЕ' (ТАРИХ-01) өткен жылы болмағандықтан бұл сөздікте
# жоқ — оның орнына шақырушы тарап (app.py) 'ТАМЫЗ' (ТАРИХ-11) деректерін
# қолданады.
PRIOR_YEAR_MONTH_TO_STREAM = {
    "ТАМЫЗ": "ТАРИХ-11",
    "ҚЫРКҮЙЕК": "ТАРИХ-21",
    "ҚАЗАН": "ТАРИХ-31",
    "ҚАРАША": "ТАРИХ-41",
    "ЖЕЛТОҚСАН": "ТАРИХ-51",
    "ҚАҢТАР": "ТАРИХ-61",
    "АҚПАН": "ТАРИХ-71",
    "НАУРЫЗ": "ТАРИХ-81",
    "СӘУІР": "ТАРИХ-91",
    "МАМЫР": "ТАРИХ-101",
}

_PRIOR_YEAR_MONTH_ROW_RE = re.compile(r"^(\d)-АЙ$")
_PRIOR_YEAR_DT_BT_HEADER_RE = re.compile(r"^([А-ЯӘҒҚҢӨҰҮҺІЁ]+)\s+ШЫҒАРМ$")


class SheetFetchError(Exception):
    pass


def _export_url(raw_url: str) -> str:
    raw_url = raw_url.strip()
    if not raw_url:
        raise SheetFetchError("Сілтеме бос болмауы керек.")
    if "format=xlsx" in raw_url:
        return raw_url
    match = SHEET_ID_RE.search(raw_url)
    if not match:
        raise SheetFetchError(
            "Google Sheets сілтемесін тани алмадым. "
            "Сілтеме '/spreadsheets/d/...' түрінде болу керек."
        )
    sheet_id = match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def fetch_workbook(raw_url: str):
    """Google Sheets кестесінің БАРЛЫҚ парақтарын (лист/tab) жүктеп қайтарады.
    Әр парақ бір куратордың кестесі деп есептеледі — парақ атауы куратор аты
    ретінде қолданылады. Қайтарады: [(sheet_name, rows), ...] тізімі."""
    export_url = _export_url(raw_url)
    try:
        with urlopen(export_url, timeout=30) as resp:
            raw_bytes = resp.read()
    except urllib.error.HTTPError as e:
        if e.code in (401, 403):
            raise SheetFetchError(
                "Кестеге қол жеткізе алмадым (рұқсат жоқ). Google Sheets-те "
                "'Файл → Ортаққа бөлу → Сілтемесі бар кез келген адам → Көруші' "
                "етіп кестені ашық етіп қойыңыз."
            ) from e
        raise SheetFetchError(f"Кестені жүктеу сәтсіз аяқталды (HTTP {e.code}).") from e
    except urllib.error.URLError as e:
        raise SheetFetchError(f"Кестені жүктеу сәтсіз аяқталды: {e.reason}") from e

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise SheetFetchError(f"Кестені оқу сәтсіз аяқталды: {e}") from e

    sheets = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v).strip() for v in row])
        while rows and all(cell == "" for cell in rows[-1]):
            rows.pop()
        if rows:
            sheets.append((ws.title, rows))

    if not sheets:
        raise SheetFetchError("Кестеде деректер табылмады.")
    return sheets


def _student_column_index(lower_header):
    idx = _find_by_keywords(lower_header, STUDENT_KEYWORDS)
    return idx if idx is not None else 0


def _strip_header_title_row(rows):
    """Кейбір парақтарда нақты баған атаулары (Оқушының аты-жөні, Балл,
    т.б.) орналасқан жолдың алдында тақырып жолы тұрады (мыс., '1-АЙ
    ДЕҢГЕЙЛІК ТЕСТ' деген жалғыз ұяшық) — оны нақты header деп қабылдап
    алсақ, содан кейінгі бағандар мүлде ескерілмей қалады. Алдыңғы 3
    жолдың ішінен толтырылған ұяшығы 2-ден кем болатындарын (тақырып жолы
    деп) өткізіп жібереміз де, содан кейінгі жолды header ретінде аламыз."""
    idx = 0
    while idx < min(len(rows) - 1, 3):
        non_empty = sum(1 for c in rows[idx] if c is not None and str(c).strip() != "")
        if non_empty >= 2:
            break
        idx += 1
    return rows[idx:]


def _numeric_cell(raw):
    """Ұяшықтан сан алады, сан шықпаса (бос, '-', мәтін) None қайтарады."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if text in ("", "-", "—"):
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def _parse_results_sheet(ws):
    """Бір парақтың жолдарын оқиды: бірінші баған оқушының аты-жөні, одан
    кейінгі бағандар — сан болса ғана балл ретінде алынады (метадеректер
    бағандары мен '-' ұяшықтар өздігінен сүзіліп қалады, себебі олардан
    сан шықпайды). Қайтарады: [(student, score), ...] — пән атауы бөлек
    (парақ атауынан) белгіленетіндіктен, мұнда жоқ."""
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    while rows and all(cell is None or str(cell).strip() == "" for cell in rows[-1]):
        rows.pop()
    rows = _strip_header_title_row(rows)
    if len(rows) < 2:
        return []

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    lower_header = [h.lower() for h in header]
    body = rows[1:]

    student_idx = _student_column_index(lower_header)

    pairs = []
    for row in body:
        student_cell = row[student_idx] if student_idx < len(row) else None
        student = str(student_cell).strip() if student_cell is not None else ""
        if not student or is_summary_row(student) or is_template_row(student):
            continue
        for i in range(len(header)):
            if i == student_idx or i >= len(row):
                continue
            score = _numeric_cell(row[i])
            if score is None:
                continue
            pairs.append((student, score))
    return pairs


# 'Шығармашылық' парағында оқушы бөлек пәндерден балл алады — Қ. Тарих
# (20 балл) және Оқу сауаттылығы (10 балл) — соларды қосқанда жалпы
# максимум 30 балл болады. Парақ атауы кейде 'Шығармашылық', кейде
# қысқартылып 'Шығарым' деп те жазылады — екеуінің де ортақ түбірі
# 'ШЫҒАР' болғандықтан сол бойынша анықтаймыз.
CREATIVE_SUBJECT_KEYWORD = "ШЫҒАР"
CREATIVE_MAX_SCORE = 30
CREATIVE_HISTORY_MAX_SCORE = 20
CREATIVE_LITERACY_MAX_SCORE = 10
# Қосынды баллдан бөлек, пән бойынша бөлек көрсету үшін (мыс. 'Ортақ
# анализде' 'Тарих пәні бойынша' / 'Оқу сауаттылығы пәні бойынша' деп) —
# осы жұрнақтармен бөлек пән ретінде сақталады.
CREATIVE_HISTORY_SUFFIX = " — Тарих"
CREATIVE_LITERACY_SUFFIX = " — Оқу сауаттылығы"

# Кейбір парақтарда әр пәннің баллынан бөлек, содан бұрын есептеп қойылған
# 'Жалпы балл' деген қорытынды баған да болады — ондай баған болса, оны
# қайта өз бетімізше қоспай (екі есе саналып кетпес үшін), тікелей сол
# бағанның мәнін оқушының қосынды баллы ретінде аламыз.
CREATIVE_TOTAL_COLUMN_KEYWORDS = ["жалпы балл", "жалпы", "қорытынды", "итого", "total", "сумма"]
CREATIVE_HISTORY_COLUMN_KEYWORDS = ["тарих"]
CREATIVE_LITERACY_COLUMN_KEYWORDS = ["оқу сауат", "сауаттылығы"]
CREATIVE_LITERACY_COLUMN_EXCLUDE_KEYWORDS = ["математик"]


def _find_creative_total_column(lower_header, student_idx):
    for i, h in enumerate(lower_header):
        if i == student_idx:
            continue
        if any(kw in h for kw in CREATIVE_TOTAL_COLUMN_KEYWORDS):
            return i
    return None


def _find_creative_subject_columns(lower_header, student_idx):
    """'Қазақстан тарихы' және 'Оқу сауаттылығы' бағандарын табады —
    'Математикалық сауаттылық' сияқты ұқсас атаулармен шатаспас үшін
    сауаттылық бағанын анықтауда 'математик' сөзі бар бағандар өткізіп
    жіберіледі."""
    history_idx = literacy_idx = None
    for i, h in enumerate(lower_header):
        if i == student_idx:
            continue
        if history_idx is None and any(kw in h for kw in CREATIVE_HISTORY_COLUMN_KEYWORDS):
            history_idx = i
        elif (
            literacy_idx is None
            and any(kw in h for kw in CREATIVE_LITERACY_COLUMN_KEYWORDS)
            and not any(kw in h for kw in CREATIVE_LITERACY_COLUMN_EXCLUDE_KEYWORDS)
        ):
            literacy_idx = i
    return history_idx, literacy_idx


def _parse_creative_results_sheet(ws):
    """'Шығармашылық' парағын оқиды: оқушы аты-жөнінен басқа әр баған
    бөлек пәннің баллы (Қ. Тарих — 20, Оқу сауаттылығы — 10). Әр оқушы
    үшін (қосынды балл, тарих баллы, сауаттылық баллы) үштігін қайтарады —
    қосынды 'Ортақ анализдегі' жалпы Шығармашылық көрсеткіші үшін (жалпы
    максимум — 30 балл), ал жеке пән баллдары пән бойынша бөлек көрсету
    үшін қолданылады. Парақта дайын 'Жалпы балл' бағаны болса, қосындыны
    қайта есептемей, сол бағанның өзі қолданылады — бірақ жеке пән
    баллдары әрдайым өз бағандарынан оқылады. Кемінде бір баған
    толтырылған оқушылар ғана қайтарылады."""
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    while rows and all(cell is None or str(cell).strip() == "" for cell in rows[-1]):
        rows.pop()
    rows = _strip_header_title_row(rows)
    if len(rows) < 2:
        return []

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    lower_header = [h.lower() for h in header]
    body = rows[1:]

    student_idx = _student_column_index(lower_header)
    total_idx = _find_creative_total_column(lower_header, student_idx)
    history_idx, literacy_idx = _find_creative_subject_columns(lower_header, student_idx)

    results = []
    for row in body:
        student_cell = row[student_idx] if student_idx < len(row) else None
        student = str(student_cell).strip() if student_cell is not None else ""
        if not student or is_summary_row(student) or is_template_row(student):
            continue

        history_score = _numeric_cell(row[history_idx]) if history_idx is not None and history_idx < len(row) else None
        literacy_score = _numeric_cell(row[literacy_idx]) if literacy_idx is not None and literacy_idx < len(row) else None

        if total_idx is not None:
            total = _numeric_cell(row[total_idx] if total_idx < len(row) else None)
        else:
            total = None
            for i in range(len(header)):
                if i == student_idx or i >= len(row):
                    continue
                score = _numeric_cell(row[i])
                if score is None:
                    continue
                total = (total or 0.0) + score

        if total is None and history_score is None and literacy_score is None:
            continue
        results.append((student, total, history_score, literacy_score))
    return results


def parse_results_file(file_bytes):
    """ДЕҢГЕЙЛІК/БАЙҚАУ ТЕСТ санатына жүктелген нәтиже файлын (.xlsx) оқиды.
    Файлда екі бөлек парақ (лист) болуы керек — 'Тарих жалпы' (әр оқушы бір
    пәннен, 20 балл) және 'Шығармашылық' ('ШЫҒАРМ' сөзі бар парақ атауы —
    әр оқушы Қ. Тарих (20) пен Оқу сауаттылығы (10) деген екі бөлек баған
    бойынша балл алады, қосындысы 30 балл, бір жолға біріктіріліп
    қайтарылады). ПӘН АТАУЫ РЕТІНДЕ СОЛ ПАРАҚТЫҢ АТАУЫ алынады. 'ҮЛГІ'
    парағы және 'Орташа балл' сияқты қорытынды жолдар өткізіп жіберіледі.
    Қайтарады: [{"student":.., "subject":.., "score":.., "max_score":..}, ...]."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise SheetFetchError(f"Файлды оқу сәтсіз аяқталды: {e}") from e

    entries = []
    for ws in wb.worksheets:
        sheet_name = (ws.title or "").strip()
        if not sheet_name or is_template_sheet(sheet_name):
            continue
        if CREATIVE_SUBJECT_KEYWORD in sheet_name.upper():
            for student, total, history_score, literacy_score in _parse_creative_results_sheet(ws):
                if total is not None:
                    entries.append({
                        "student": student, "subject": sheet_name,
                        "score": total, "max_score": CREATIVE_MAX_SCORE,
                    })
                if history_score is not None:
                    entries.append({
                        "student": student, "subject": sheet_name + CREATIVE_HISTORY_SUFFIX,
                        "score": history_score, "max_score": CREATIVE_HISTORY_MAX_SCORE,
                    })
                if literacy_score is not None:
                    entries.append({
                        "student": student, "subject": sheet_name + CREATIVE_LITERACY_SUFFIX,
                        "score": literacy_score, "max_score": CREATIVE_LITERACY_MAX_SCORE,
                    })
        else:
            for student, score in _parse_results_sheet(ws):
                entries.append({"student": student, "subject": sheet_name, "score": score})

    if not entries:
        raise SheetFetchError("Файлдан бірде-бір дұрыс нәтиже табылмады.")
    return entries


def rows_to_dicts(rows):
    header = [h.strip() for h in rows[0]]
    body = rows[1:]
    return header, body


STUDENT_KEYWORDS = ["аты-жөні", "оқушы", "фио", "тегі", "аты", "name", "student"]
SCORE_KEYWORDS = ["балл", "ұпай", "score", "нәтиже", "ст"]
MAX_SCORE_KEYWORDS = ["максимум", "максимал", "жалпы балл", "max"]
SUBJECT_KEYWORDS = ["пән", "предмет", "subject"]
TOPIC_KEYWORDS = ["тақырып", "тема", "topic"]
ROW_NUMBER_HEADERS = {"№", "n", "no", "нөмір", "рет саны", "п/п", "р/с"}
NUMERIC_FALLBACK_MIN_RATIO = 0.7
NUMERIC_SAMPLE_ROWS = 50

SUMMARY_ROW_KEYWORDS = [
    "орташа",
    "орта балл",
    "ортақ балл",
    "ортақ ұпай",
    "жиынтық",
    "қорытынды",
    "барлығы",
    "барлыгы",
    "итого",
    "всего",
    "average",
    "total",
    "сумма",
]


def is_summary_row(name: str) -> bool:
    """Кестенің соңындағы 'Орташа ұпайы', 'Барлығы' сияқты қорытынды жолдарды
    (нақты оқушы емес) анықтайды — олар импорттан тыс қалдырылады."""
    lower = (name or "").strip().lower()
    if not lower:
        return False
    return any(kw in lower for kw in SUMMARY_ROW_KEYWORDS)


_TEMPLATE_WORD_RE = re.compile(r"[a-zа-яёіңғүұқөһ]+", re.IGNORECASE)


def _has_template_word(text) -> bool:
    """Мәтінде 'үлгі' деген жеке сөз бар-жоғын тексереді (толық сөз ретінде
    ғана — 'Үлгібек' сияқты нақты есімдермен шатаспас үшін жай substring
    емес, сөз шекарасы бойынша салыстырады)."""
    if not text:
        return False
    return "үлгі" in _TEMPLATE_WORD_RE.findall(str(text).lower())


def is_template_sheet(name: str) -> bool:
    """Кураторларға парақты қалай толтыру керектігін көрсету үшін қалдырылған
    'ҮЛГІ' атты парақты (нақты куратор емес) анықтайды — импортқа қосылмайды.
    Дәл 'ҮЛГІ' атаумен қатар, 'ҮЛГІ 1', 'ҮЛГІ (мысал)' сияқты құрамында
    'үлгі' сөзі бар атауларды да таниды."""
    return _has_template_word(name)


def is_template_row(name: str) -> bool:
    """Парақ ІШІНДЕ, жеке жол/оқушы ретінде қалдырылған 'ҮЛГІ' деп
    белгіленген мысал жазбаны анықтайды (нақты оқушы/куратор емес) —
    is_template_sheet парақтың ТҮГЕЛ атауын тексерсе, бұл нақты жол
    ішіндегі оқушы/куратор атын тексереді."""
    return _has_template_word(name)


def _find_by_keywords(lower_header, keywords, exclude=()):
    for i, h in enumerate(lower_header):
        if i in exclude:
            continue
        if h in keywords:
            return i
    for i, h in enumerate(lower_header):
        if i in exclude:
            continue
        if any(kw in h for kw in keywords):
            return i
    return None


def _is_sequential_index(values):
    """Баған мәндері рет нөмірі (1, 2, 3, ...) ме — тексереді, топ ортасында
    (мыс., бір парақта бірнеше сынып біріктірілгенде) 1-ден қайта басталуы
    мүмкін екенін де ескереді. Кестенің бірінші бағаны көбіне '№'/'рет саны'
    болғанымен, кейде тақырыпсыз (бос атаумен) келеді, сондықтан баған
    атауына қарамай, мәндерінің өзінен де осындай рет нөмірі екенін
    анықтаймыз (әйтпесе ол баллдың орнына қате таңдалып қалуы мүмкін)."""
    if len(values) < 5:
        return False
    try:
        nums = [float(v.replace(",", ".")) for v in values]
    except ValueError:
        return False
    if any(n != round(n) for n in nums):
        return False
    if nums[0] != 1:
        return False
    for prev, cur in zip(nums, nums[1:]):
        if cur != prev + 1 and cur != 1:
            return False
    return True


def _find_numeric_column(header, body, exclude):
    lower_header = [h.lower().strip() for h in header]
    best_idx, best_ratio = None, 0.0
    for i in range(len(header)):
        if i in exclude or lower_header[i] in ROW_NUMBER_HEADERS:
            continue
        numeric = 0
        total = 0
        values = []
        for row in body[:NUMERIC_SAMPLE_ROWS]:
            if i >= len(row) or row[i] == "":
                continue
            total += 1
            try:
                float(row[i].replace(",", "."))
                numeric += 1
                values.append(row[i])
            except ValueError:
                pass
        if total == 0 or _is_sequential_index(values):
            continue
        ratio = numeric / total
        if ratio > best_ratio:
            best_ratio, best_idx = ratio, i
    if best_ratio >= NUMERIC_FALLBACK_MIN_RATIO:
        return best_idx
    return None


def guess_columns(header, body):
    """Баған атаулары бойынша (немесе сан басым баған эвристикасымен) оқушы,
    балл, пән, тақырып, максимум балл бағандарын автоматты анықтайды —
    пайдаланушыдан қолмен таңдауды талап етпейді."""
    lower = [h.lower().strip() for h in header]

    student_idx = _find_by_keywords(lower, STUDENT_KEYWORDS)
    exclude = {student_idx} if student_idx is not None else set()

    score_idx = _find_by_keywords(lower, SCORE_KEYWORDS, exclude=exclude)
    if score_idx is None:
        score_idx = _find_numeric_column(header, body, exclude=exclude)
    if score_idx is not None:
        exclude = exclude | {score_idx}

    max_score_idx = _find_by_keywords(lower, MAX_SCORE_KEYWORDS, exclude=exclude)
    subject_idx = _find_by_keywords(lower, SUBJECT_KEYWORDS, exclude=exclude)
    topic_idx = _find_by_keywords(lower, TOPIC_KEYWORDS, exclude=exclude)

    return {
        "student": student_idx,
        "score": score_idx,
        "max_score": max_score_idx,
        "subject": subject_idx,
        "topic": topic_idx,
    }


def _parse_prior_year_number(raw):
    text = (raw or "").strip()
    if text in ("", "-", "-%"):
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def _parse_prior_year_st_sheet(rows):
    """'СТ нәтижелері' парағын оқиды: әр блок бір поток (ай атымен
    белгіленген), одан кейінгі 'N-АЙ' жолдарының 6-бағаны (ОРТАҚ) — сол
    айдың орташа баллы. Қайтарады: [(stream_code, month_number, avg), ...]."""
    out = []
    i = 0
    while i < len(rows):
        row = rows[i]
        col0 = (row[0] or "").strip()
        stream_code = PRIOR_YEAR_MONTH_TO_STREAM.get(col0)
        is_header = stream_code and len(row) > 1 and (row[1] or "").strip() == "1-АПТА"
        if is_header:
            i += 1
            while i < len(rows):
                r = rows[i]
                m = _PRIOR_YEAR_MONTH_ROW_RE.match((r[0] or "").strip())
                if not m:
                    break
                month_number = int(m.group(1))
                avg = _parse_prior_year_number(r[5]) if len(r) > 5 else None
                out.append((stream_code, month_number, avg))
                i += 1
        else:
            i += 1
    return out


def _parse_prior_year_dt_bt_sheet(rows):
    """'ДТ нәтижелері'/'БТ нәтижелері' парағын оқиды: әр блок бір поток
    ('<АЙ АТЫ> ШЫҒАРМ' деп басталады), одан кейінгі 'N-АЙ' жолдарының
    2-бағаны (ОРТАҚ БАЛЛ) — сол айдың орташа баллы. Қайтарады:
    [(stream_code, month_number, avg), ...]."""
    out = []
    i = 0
    while i < len(rows):
        row = rows[i]
        col0 = (row[0] or "").strip()
        header_match = _PRIOR_YEAR_DT_BT_HEADER_RE.match(col0)
        stream_code = (
            PRIOR_YEAR_MONTH_TO_STREAM.get(header_match.group(1)) if header_match else None
        )
        if stream_code:
            i += 1
            while i < len(rows):
                r = rows[i]
                m = _PRIOR_YEAR_MONTH_ROW_RE.match((r[0] or "").strip())
                if not m:
                    break
                month_number = int(m.group(1))
                avg = _parse_prior_year_number(r[1]) if len(r) > 1 else None
                out.append((stream_code, month_number, avg))
                i += 1
        else:
            i += 1
    return out


def parse_prior_year_report(raw_url):
    """Жылдық отчет Google Sheets кестесін оқып, СТ/ДТ/БТ парақтарынан
    (АТ-ны қоспай — ол бөлек есептеледі/жоқ) поток пен ай бойынша орташа
    баллды алады. Қайтарады: [(category, stream_code, month_number, avg), ...]
    мұндағы category — 'sabaq_tapsyru', 'dt' немесе 'bt'."""
    sheets = fetch_workbook(raw_url)
    by_name = {name.strip(): rows for name, rows in sheets}

    out = []
    if "СТ нәтижелері" in by_name:
        for stream_code, month_number, avg in _parse_prior_year_st_sheet(by_name["СТ нәтижелері"]):
            out.append(("sabaq_tapsyru", stream_code, month_number, avg))
    if "ДТ нәтижелері" in by_name:
        for stream_code, month_number, avg in _parse_prior_year_dt_bt_sheet(by_name["ДТ нәтижелері"]):
            out.append(("dt", stream_code, month_number, avg))
    if "БТ нәтижелері" in by_name:
        for stream_code, month_number, avg in _parse_prior_year_dt_bt_sheet(by_name["БТ нәтижелері"]):
            out.append(("bt", stream_code, month_number, avg))

    if not out:
        raise SheetFetchError("Жылдық отчеттен СТ/ДТ/БТ парақтары табылмады.")
    return out


LS_SHEET_NAME = "ТАРИХ"
_LS_WEEK_LABEL_RE = re.compile(r"^\d+-АЙ \d+-АПТА$")
_LS_REQUIRED_COLUMNS = ["күні", "мұғалім", "поток", "ұнау пайызы", "қатысым пайызы"]

# Live сабақ импорты бағдарлама бойынша бөлек жүреді — Smart-тың потоктары
# 'ТАРИХ-XX', Junior-дікі 'JUNIOR-XX' болатын бекітілген атаумен сәйкес
# келу үшін (streams кестесіндегі кодтармен бірдей, DEFAULT_PROGRAMS-те
# анықталған).
LS_STREAM_PREFIX_BY_PROGRAM = {"smart": "ТАРИХ", "junior": "JUNIOR"}

# Junior-дың кейбір потоктары LS кестесінің 'поток' бағанында сандық код
# емес, брендтік атаумен жазылады (мыс. 'ZEREK' — JUNIOR-01, 'USHQYN' —
# JUNIOR-11 потогінің баламасы, екеуі бір поток). Тек Junior бағдарламасына
# қатысты — Smart-та мұндай балама атаулар жоқ.
LS_JUNIOR_STREAM_NAME_ALIASES = {"ZEREK": "JUNIOR-01", "USHQYN": "JUNIOR-11"}
# Ағым кодынан LS интерфейсінде көрсетілетін атауға (керісінше бағыт) —
# осы потоктар LS бетінде де кестедегідей 'ZEREK'/'USHQYN' болып көрінуі үшін.
LS_STREAM_DISPLAY_NAMES = {code: name for name, code in LS_JUNIOR_STREAM_NAME_ALIASES.items()}


def _find_column_by_name(header, name):
    target = name.strip().lower()
    for i, h in enumerate(header):
        if (h or "").strip().lower() == target:
            return i
    return None


def _find_ls_worksheet(wb):
    """Керек 5 бағаны (күні/мұғалім/поток/ҰНАУ ПАЙЫЗЫ/қатысым пайызы) бар
    парақты іздейді — алдымен '{LS_SHEET_NAME}' атымен, таппаса кестенің
    басқа парақтарынан да іздейді (Junior кестесінде парақ басқаша
    аталуы мүмкін болғандықтан)."""
    candidates = [LS_SHEET_NAME] + [n for n in wb.sheetnames if n != LS_SHEET_NAME]
    for name in candidates:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            continue
        header = [str(h).strip().lower() if h is not None else "" for h in first_row]
        if all(col in header for col in _LS_REQUIRED_COLUMNS):
            return name
    return None


def _normalize_ls_stream_code(raw_potok, program):
    """'01'/'11'/11.0 сияқты 'поток' мәнінен, бағдарламаға сай 'ТАРИХ-01'
    немесе 'JUNIOR-01' сияқты толық ағым кодын жасайды. Junior-да 'ZEREK'/
    'USHQYN' сияқты брендтік атаулар да LS_JUNIOR_STREAM_NAME_ALIASES
    арқылы тиісті ағым кодына сәйкестендіріледі."""
    if raw_potok is None:
        return None
    text = str(raw_potok).strip()
    if not text:
        return None
    if program == "junior":
        alias = LS_JUNIOR_STREAM_NAME_ALIASES.get(text.upper())
        if alias:
            return alias
    try:
        num = int(float(text))
    except ValueError:
        return None
    prefix = LS_STREAM_PREFIX_BY_PROGRAM[program]
    return f"{prefix}-{num:02d}"


def parse_ls_report(raw_url, program):
    """Live сабақ бағалау кестесінің тиісті парағын оқиды (алдымен 'ТАРИХ'
    атымен, таппаса керек бағандары бар кез келген парақтан). Парақ
    жол-жол 'N-АЙ M-АПТА' деген жалғыз ұяшықты апта белгісімен бөлінген
    (одан кейінгі session жолдарының бәрі сол аптаға жатады, келесі
    белгі жолы кездескенше). Әр session жолынан күні/мұғалім/поток/
    ҰНАУ ПАЙЫЗЫ/қатысым пайызы алынады, 'поток' ('01'/'11') program-ға
    сай ('smart' → 'ТАРИХ-01', 'junior' → 'JUNIOR-01') толық ағым
    кодына айналдырылады. Қайтарады: [{"session_date":.., "teacher_name":..,
    "stream_code":.., "week_label":.., "like_percent":.., "attendance_percent":..}, ...]."""
    export_url = _export_url(raw_url)
    try:
        with urlopen(export_url, timeout=30) as resp:
            raw_bytes = resp.read()
    except urllib.error.HTTPError as e:
        if e.code in (401, 403):
            raise SheetFetchError(
                "Кестеге қол жеткізе алмадым (рұқсат жоқ). Google Sheets-те "
                "'Файл → Ортаққа бөлу → Сілтемесі бар кез келген адам → Көруші' "
                "етіп кестені ашық етіп қойыңыз."
            ) from e
        raise SheetFetchError(f"Кестені жүктеу сәтсіз аяқталды (HTTP {e.code}).") from e
    except urllib.error.URLError as e:
        raise SheetFetchError(f"Кестені жүктеу сәтсіз аяқталды: {e.reason}") from e

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise SheetFetchError(f"Кестені оқу сәтсіз аяқталды: {e}") from e

    sheet_name = _find_ls_worksheet(wb)
    if sheet_name is None:
        raise SheetFetchError(
            "Кестеден керек бағандары бар парақ табылмады "
            "(күні / мұғалім / поток / ҰНАУ ПАЙЫЗЫ / қатысым пайызы)."
        )
    ws = wb[sheet_name]

    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    if not rows:
        raise SheetFetchError(f"'{sheet_name}' парағы бос.")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    body = rows[1:]

    date_idx = _find_column_by_name(header, "күні")
    teacher_idx = _find_column_by_name(header, "мұғалім")
    stream_idx = _find_column_by_name(header, "поток")
    like_idx = _find_column_by_name(header, "ҰНАУ ПАЙЫЗЫ")
    attendance_idx = _find_column_by_name(header, "қатысым пайызы")
    if None in (date_idx, teacher_idx, stream_idx, like_idx, attendance_idx):
        raise SheetFetchError(
            f"'{sheet_name}' парағында керек бағандар табылмады "
            "(күні / мұғалім / поток / ҰНАУ ПАЙЫЗЫ / қатысым пайызы)."
        )

    entries = []
    current_week_label = None
    for row in body:
        first_cell = row[0] if row else None
        teacher = row[teacher_idx] if teacher_idx < len(row) else None
        potok = row[stream_idx] if stream_idx < len(row) else None
        kuni = row[date_idx] if date_idx < len(row) else None

        if teacher is None and potok is None and kuni is None:
            if isinstance(first_cell, str) and _LS_WEEK_LABEL_RE.match(first_cell.strip()):
                current_week_label = first_cell.strip()
            continue

        teacher_name = str(teacher).strip() if teacher is not None else ""
        if not teacher_name:
            continue

        stream_code = _normalize_ls_stream_code(potok, program)
        if not stream_code:
            continue

        like_score = _numeric_cell(row[like_idx] if like_idx < len(row) else None)
        attendance_score = _numeric_cell(row[attendance_idx] if attendance_idx < len(row) else None)

        # datetime мәнінің уақыт бөлігін де сақтаймыз (тек күнін алмай) —
        # бір күнде бірнеше session болса, солардың реті (уақыты бойынша)
        # осыдан кейін дұрыс сұрыпталуы үшін қажет.
        if hasattr(kuni, "isoformat"):
            session_date = kuni.isoformat()
        elif kuni is not None:
            session_date = str(kuni).strip()
        else:
            session_date = None

        entries.append({
            "session_date": session_date,
            "teacher_name": teacher_name,
            "stream_code": stream_code,
            "week_label": current_week_label,
            "like_percent": round(like_score * 100, 1) if like_score is not None else None,
            "attendance_percent": round(attendance_score * 100, 1) if attendance_score is not None else None,
        })

    if not entries:
        raise SheetFetchError(f"'{sheet_name}' парағынан бірде-бір дұрыс жол табылмады.")
    return entries
